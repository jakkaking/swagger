import os
import shutil
import tempfile
from datetime import datetime
from typing import Optional

import psycopg2
from dotenv import load_dotenv
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook

# Load environment variables from .env file
load_dotenv()

router = APIRouter(prefix="/mappings", tags=["Mappings"])

# Pydantic model for input JSON body
class MappingInput(BaseModel):
    schema: str
    table: str
    author: str
    load_strategy: str
    ilm_strategy: str
    pi_cols: Optional[str] = ""
    pk_cols: Optional[str] = ""
    fk_cols: Optional[str] = ""
    compression_cols: Optional[str] = ""
    unicode_cols: Optional[str] = ""
    pii_cols: Optional[str] = ""

def get_gp_connection():
    try:
        conn = psycopg2.connect(
            dbname=os.getenv("GP_DB"),
            user=os.getenv("GP_USER"),
            password=os.getenv("GP_PASSWORD"),
            host=os.getenv("GP_HOST"),
            port=os.getenv("GP_PORT"),
        )
        return conn
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to connect to Greenplum: {e}")

def fetch_table_metadata(conn, schema, table):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT 
                column_name, 
                udt_name, 
                character_maximum_length,
                numeric_precision,
                numeric_scale,
                is_nullable
            FROM information_schema.columns
            WHERE table_schema = %s AND table_name = %s
            ORDER BY ordinal_position
        """, (schema, table))
        return cur.fetchall()

def format_data_type(udt_name, char_len, num_prec, num_scale):
    udt_name = udt_name.lower()
    if udt_name in ['varchar']:
        return f"{udt_name}({char_len})" if char_len else udt_name
    elif udt_name in ['char', 'bpchar']:
        return f"CHAR({char_len})" if char_len else udt_name
    elif udt_name in ['timestamp']:
        return f"{udt_name}(6)"
    elif udt_name in ['numeric', 'decimal']:
        if num_prec:
            return "DECIMAL(18,0)"
        else:
            return udt_name
    elif udt_name in ['int4']:
        return "INTEGER"
    elif udt_name in ['int2']:
        return "SMALLINT"
    elif udt_name in ['int8']:
        return "BIGINT"
    elif udt_name in ['timestamptz']:
        return "TIMESTAMP(6) WITH TIME ZONE"
    else:
        return udt_name

def parse_csv(s: Optional[str]):
    if not s:
        return []
    return [col.strip() for col in s.split(",") if col.strip()]

def generate_mapping_excel(template_path, output_path, schema, table, metadata, author, load_sta, ILM_str,
                           pi_cols, pk_cols, fk_cols, comp_cols, unicode_cols, pii_cols):
    # Copy template to output
    shutil.copy(template_path, output_path)

    wb = load_workbook(output_path)
    ws_mapping = wb["Data Mapping"]
    ws_info = wb["Mapping Information"]

    headers = [cell.value.strip() if cell.value else "" for cell in ws_mapping[1]]
    col_indexes = {h: i + 1 for i, h in enumerate(headers)}

    for i, (col_name, udt_name, char_len, num_prec, num_scale, nullable) in enumerate(metadata, start=2):
        datatype = format_data_type(udt_name, char_len, num_prec, num_scale).upper()

        ws_mapping.cell(row=i, column=col_indexes["Seq#"], value=i - 1)
        ws_mapping.cell(row=i, column=col_indexes["Target Column Name"], value=col_name)
        ws_mapping.cell(row=i, column=col_indexes["Target Datatype"], value=datatype)
        ws_mapping.cell(row=i, column=col_indexes["Nullable \n(Y/N)"], value="Y" if nullable == "YES" else "N")

        lower_col = col_name.lower()
        if any(
            kw in lower_col
            for kw in [
                "op_type",
                "pos",
                "op_ts",
                "src_sys_nm",
                "kfk_ins_dtsz",
                "dw_row_hash_val",
                "dw_src_site_id",
                "dw_ins_dtsz",
                "dw_upd_dtsz",
                "dw_ld_grp_val",
                "dw_etl_sess_nm",
            ]
        ):
            ws_mapping.cell(row=i, column=col_indexes["Source Column Name"], value="ETL Derived")
            ws_mapping.cell(row=i, column=col_indexes["Source Data type"], value="ETL Derived")
            ws_mapping.cell(row=i, column=col_indexes["Source Schema"], value="ETL Derived")
            ws_mapping.cell(row=i, column=col_indexes["Source Table Name"], value="ETL Derived")
        else:
            ws_mapping.cell(row=i, column=col_indexes["Source Column Name"], value=col_name)
            ws_mapping.cell(row=i, column=col_indexes["Source Data type"], value=datatype)
            ws_mapping.cell(row=i, column=col_indexes["Source Schema"], value=schema)
            ws_mapping.cell(row=i, column=col_indexes["Source Table Name"], value=table)

        # Transform Comments based on column name
        match lower_col:
            case "op_type":
                comment = "I/U/D"
            case "pos":
                comment = "Position"
            case "op_ts":
                comment = "Timestamp"
            case "src_sys_nm":
                comment = "GTM"
            case "kfk_ins_dtsz":
                comment = "kafka timestamp"
            case "dw_row_hash_val":
                comment = "12345"
            case "dw_src_site_id":
                comment = "4101"
            case "dw_ins_dtsz":
                comment = "current_timestamp"
            case "dw_upd_dtsz":
                comment = "current_timestamp"
            case "dw_ld_grp_val":
                comment = "123456"
            case "dw_etl_sess_nm":
                comment = "ETL/SS/GPSS"
            case _:
                comment = "Straight Pull"

        ws_mapping.cell(row=i, column=col_indexes["Transform Comments"], value=comment)
        ws_mapping.cell(row=i, column=col_indexes["Mod Date"], value="")
        ws_mapping.cell(row=i, column=col_indexes["Target Column Description"], value=col_name.replace("_", " "))

        # Y/N flags
        ws_mapping.cell(row=i, column=col_indexes["PI/DK \n(Y/N)"], value="Y" if col_name in pi_cols else "N")
        ws_mapping.cell(row=i, column=col_indexes["PK\n(Y/N)"], value="Y" if col_name in pk_cols else "N")
        ws_mapping.cell(row=i, column=col_indexes["FK\n(Y/N)"], value="Y" if col_name in fk_cols else "N")
        ws_mapping.cell(row=i, column=col_indexes["Comp-\nression\n(Y/N)"], value="Y" if col_name in comp_cols else "N")
        ws_mapping.cell(row=i, column=col_indexes["Uni-code\n(Y/N)"], value="Y" if col_name in unicode_cols else "N")
        ws_mapping.cell(row=i, column=col_indexes["PII\n(Y/N)"], value="Y" if col_name in pii_cols else "N")
        ws_mapping.cell(row=i, column=col_indexes["Security Classification"], value="internal")

    # Mapping Information sheet
    headers_info = [cell.value.strip() if cell.value else "" for cell in ws_info[1]]
    col_info = {h: i + 1 for i, h in enumerate(headers_info)}

    ws_info.cell(row=2, column=col_info["Sno"], value=1)
    ws_info.cell(row=2, column=col_info["Domain"], value="GOSC")
    ws_info.cell(row=2, column=col_info["Project ID - Project Name"], value="DSC Logistics/Trade")
    ws_info.cell(row=2, column=col_info["Mapping Version"], value="1.0")
    ws_info.cell(row=2, column=col_info["Additional Information"], value="")
    ws_info.cell(row=2, column=col_info["Created by Data Architect"], value=author)
    ws_info.cell(row=2, column=col_info["Created Date"], value=datetime.now().strftime("%Y-%m-%d"))
    ws_info.cell(row=2, column=col_info["Load Strategy"], value=load_sta)
    ws_info.cell(row=2, column=col_info["Data Expectations"], value="")
    ws_info.cell(row=2, column=col_info["ILM Strategy"], value=ILM_str)

    wb.save(output_path)


@router.post("/generate/", response_class=FileResponse)
async def generate_mapping(data: MappingInput):
    conn = get_gp_connection()
    tmpdir = tempfile.mkdtemp()  # manually create temp dir
    try:
        metadata = fetch_table_metadata(conn, data.schema, data.table)

        template_path = "mapping.xlsx"  # Make sure this path is correct

        output_filename = f"{data.schema.upper()}_T_{data.table.upper()}.xlsx"
        output_path = os.path.join(tmpdir, output_filename)

        generate_mapping_excel(
            template_path,
            output_path,
            data.schema,
            data.table,
            metadata,
            data.author,
            data.load_strategy,
            data.ilm_strategy,
            parse_csv(data.pi_cols),
            parse_csv(data.pk_cols),
            parse_csv(data.fk_cols),
            parse_csv(data.compression_cols),
            parse_csv(data.unicode_cols),
            parse_csv(data.pii_cols),
        )

        if not os.path.exists(output_path):
            raise HTTPException(status_code=500, detail="Failed to generate the Excel file.")

        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=output_filename,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
